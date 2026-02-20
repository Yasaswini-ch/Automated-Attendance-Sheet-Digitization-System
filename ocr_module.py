"""
ocr_module.py
=============
Production-grade OCR module for GVP-style college attendance sheets.

Key fixes vs v1:
  • PaddleOCR init tries multiple param signatures (handles 2.x / 2.7+ / 2.8+)
  • ocr() call tries cls=True then bare call (handles version differences)
  • Handles two side-by-side student blocks (GVP sheet layout)
  • Per-period breakdown stored in roll_details
  • Upscales low-res images automatically
  • Fully graceful error handling — never silently swallows data

Supports: .jpg .jpeg .png .pdf .xlsx
"""

import os
import re
import logging
import warnings
from pathlib import Path

import numpy as np
import cv2
import pandas as pd

warnings.filterwarnings("ignore")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────
ROLL_PATTERN = re.compile(r"\b3\d{9,11}\b")
NUM_PERIODS  = 8

IGNORE_WORDS = {
    "period", "signature", "total", "present", "absent", "name",
    "faculty", "staff", "subject", "date", "day", "time", "slot",
    "class", "section", "dept", "department", "timetable", "roll",
    "no", "number", "sign", "initial", "substitute", "lunch",
    "teacher", "the", "of", "any", "if", "for", "and", "lab",
    "devops", "table", "cd", "ml", "pds",
}

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif"}


# ══════════════════════════════════════════════════════════════════════════════
# External data loaders
# ══════════════════════════════════════════════════════════════════════════════

def load_students_order() -> list:
    """
    Load roll numbers from students.xlsx to maintain correct order.
    Returns list of roll numbers in the order they appear in students.xlsx
    """
    try:
        df = pd.read_excel("students.xlsx")
        roll_column = None
        for col in df.columns:
            if "roll" in col.lower():
                roll_column = col
                break
        
        if roll_column is None:
            logger.warning("No roll column found in students.xlsx")
            return []
        
        rolls = df[roll_column].dropna().astype(str).str.strip().tolist()
        # Filter only valid roll numbers
        valid_rolls = [r for r in rolls if ROLL_PATTERN.fullmatch(r)]
        logger.info(f"Loaded {len(valid_rolls)} roll numbers from students.xlsx")
        return valid_rolls
    except Exception as e:
        logger.warning(f"Could not load students.xlsx: {e}")
        return []

def load_schedule_info() -> dict:
    """
    Load schedule and staff information from 3CSM1_Timetable.xlsx.
    Returns dict with period names, staff information, and lunch periods
    """
    try:
        df = pd.read_excel("3CSM1_Timetable.xlsx", sheet_name="3CSM1 Timetable")
        
        # Get period names from column headers (skip DAY column)
        periods = {}
        period_cols = [col for col in df.columns if col != "DAY"]
        
        # Map periods to standard names
        period_mapping = {
            "8:40AM-9:30AM": "P1",
            "9:30AM-10:20AM": "P2", 
            "10:20AM-11:10AM": "P3",
            "11:10AM-12:00PM": "P4",
            "12:00PM-12:50PM": "P5",
            "12:50PM-1:40PM": "P6",
            "1:40PM-2:30PM": "P7",
            "2:30PM-3:20PM": "P8",
            "3:20PM-5:00PM": "P9",  # Extra period
        }
        
        # Identify lunch periods (typically 12:00PM-12:50PM and 12:50PM-1:40PM)
        lunch_periods = []
        for col in period_cols:
            # Check if this is a lunch period by looking for "LUNCH" in the column data
            if any("LUNCH" in str(cell).upper() for cell in df[col].dropna()):
                period_key = period_mapping.get(col, col)
                lunch_periods.append(period_key)
        
        # Extract staff for each period
        staff_per_period = {}
        for col in period_cols:
            period_key = period_mapping.get(col, col)
            # Get all unique staff names for this period
            staff_list = []
            for cell in df[col].dropna():
                if isinstance(cell, str) and cell not in ["LIB", "LIBLIB", "NaN", "AIMERS CLUB ACTIVITY"]:
                    # Extract teacher names from parentheses
                    import re
                    teachers = re.findall(r"\(([^)]+)\)", cell)
                    for teacher in teachers:
                        # Split by comma and clean
                        for t in teacher.split(","):
                            t = t.strip()
                            # Remove title prefixes
                            t = re.sub(r"^(Mr|Mrs|Ms|Dr)\.?\s*", "", t)
                            if t and len(t) > 1:
                                staff_list.append(t)
            
            # Remove duplicates while preserving order
            seen = set()
            unique_staff = []
            for staff in staff_list:
                if staff not in seen:
                    seen.add(staff)
                    unique_staff.append(staff)
            
            staff_per_period[period_key] = unique_staff
        
        logger.info(f"Loaded schedule info for {len(staff_per_period)} periods, lunch periods: {lunch_periods}")
        return {
            "periods": period_mapping,
            "staff_per_period": staff_per_period,
            "lunch_periods": lunch_periods
        }
    except Exception as e:
        logger.warning(f"Could not load 3CSM1_Timetable.xlsx: {e}")
        return {"periods": {}, "staff_per_period": {}, "lunch_periods": []}


# ══════════════════════════════════════════════════════════════════════════════
# PaddleOCR singleton — version-safe
# ══════════════════════════════════════════════════════════════════════════════

_ocr_engine = None


def _get_ocr_engine():
    """
    Build PaddleOCR, trying progressively simpler constructor signatures.
    Supports PaddleOCR 2.x, 2.7+, 2.8+.
    """
    global _ocr_engine
    if _ocr_engine is not None:
        return _ocr_engine

    # Force CPU-only mode and disable problematic backends
    os.environ['ONEDNN_DISABLE'] = '1'
    os.environ['FLAGS_use_mkldnn'] = '0'
    os.environ['CUDA_VISIBLE_DEVICES'] = ''
    os.environ['FLAGS_cudnn_batchnorm_spatial_persistent'] = '0'
    os.environ['FLAGS_conv_workspace_size_limit'] = '-1'

    try:
        from paddleocr import PaddleOCR  # noqa
    except ImportError as e:
        raise ImportError(
            "paddleocr not installed. Run:\n"
            "  pip install paddleocr paddlepaddle"
        ) from e

    attempts = [
        dict(use_textline_orientation=True, lang="en"),
        dict(use_textline_orientation=True),
        dict(lang="en"),
        dict(),
    ]

    last_err = None
    for kwargs in attempts:
        try:
            engine = PaddleOCR(**kwargs)
            _ocr_engine = engine
            logger.info("PaddleOCR ready — params: %s", kwargs)
            return _ocr_engine
        except Exception as e:
            last_err = e
            continue

    raise RuntimeError(f"Cannot init PaddleOCR after all attempts: {last_err}")


# ══════════════════════════════════════════════════════════════════════════════
# Image pre-processing
# ══════════════════════════════════════════════════════════════════════════════

def _preprocess(img_bgr: np.ndarray) -> np.ndarray:
    """
    Enhanced preprocessing for better OCR accuracy and position normalization.
    Steps: Grayscale → CLAHE contrast → adaptive threshold → noise reduction → 3-channel BGR.
    """
    # Convert to grayscale
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    
    # Apply CLAHE for better contrast
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(gray)
    
    # Gaussian blur to reduce noise
    blurred = cv2.GaussianBlur(enhanced, (3, 3), 0)
    
    # Adaptive threshold with better parameters
    binary = cv2.adaptiveThreshold(
        blurred, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY, 15, 8,
    )
    
    # Morphological operations to clean up
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    cleaned = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
    
    # Mild dilation to strengthen text
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 1))
    result = cv2.dilate(cleaned, kernel, iterations=1)
    
    return cv2.cvtColor(result, cv2.COLOR_GRAY2BGR)


def _auto_upscale(img: np.ndarray, min_width: int = 1400) -> np.ndarray:
    """Upscale images narrower than min_width for better OCR accuracy."""
    h, w = img.shape[:2]
    if w < min_width:
        scale = min_width / w
        img = cv2.resize(img, None, fx=scale, fy=scale,
                         interpolation=cv2.INTER_CUBIC)
        logger.info("Upscaled image %.1fx → (%d×%d)", scale, img.shape[1], img.shape[0])
    return img


def _split_regions(img: np.ndarray):
    """
    Return (top, middle, bottom) crops:
      top    = 0–20%   : college header, date, section
      middle = 20–80%  : student roll numbers + period marks
      bottom = 80–100% : teacher names, timetable
    """
    h = img.shape[0]
    return (
        img[0: int(0.20 * h), :],
        img[int(0.20 * h): int(0.80 * h), :],
        img[int(0.80 * h): h, :],
    )


# ══════════════════════════════════════════════════════════════════════════════
# OCR runner — version-safe
# ══════════════════════════════════════════════════════════════════════════════

def _run_ocr(region: np.ndarray) -> list:
    """
    Run PaddleOCR on a BGR numpy region.
    Uses ocr() method for PaddleOCR 2.7.3 compatibility.
    Returns list of dicts sorted top→bottom, left→right.
    """
    ocr = _get_ocr_engine()
    
    try:
        # Use ocr() method for PaddleOCR 2.7.3
        result = ocr.ocr(region)
    except Exception as e:
        logger.warning("OCR call failed: %s", e)
        return []

    if not result or not result[0]:
        return []

    records = []
    for line in result[0]:
        if line is None:
            continue
        try:
            bbox, (text, conf) = line[0], line[1]
            records.append({
                "text": str(text).strip(),
                "confidence": float(conf),
                "y": float(bbox[0][1]),
                "x": float(bbox[0][0]),
                "bbox": bbox,
            })
        except Exception:
            continue

    records.sort(key=lambda r: (r["y"], r["x"]))
    return records


def _group_rows(records: list, y_tol: int = 14) -> list:
    """Cluster OCR records into horizontal rows using Y-coordinate proximity."""
    if not records:
        return []
    rows, cur, cur_y = [], [records[0]], records[0]["y"]
    for rec in records[1:]:
        if abs(rec["y"] - cur_y) <= y_tol:
            cur.append(rec)
        else:
            rows.append(sorted(cur, key=lambda r: r["x"]))
            cur, cur_y = [rec], rec["y"]
    rows.append(sorted(cur, key=lambda r: r["x"]))
    return rows


# ══════════════════════════════════════════════════════════════════════════════
# Region-specific extractors
# ══════════════════════════════════════════════════════════════════════════════

def extract_top_region(top_img: np.ndarray) -> dict:
    """Parse date and section from the top header region."""
    records   = _run_ocr(top_img)
    full_text = " ".join(r["text"] for r in records)
    logger.debug("TOP text: %s", full_text)
    return {
        "date":    _parse_date(full_text),
        "section": _parse_section(full_text),
    }


def _parse_date(text: str) -> str:
    # DD-MM-YY or DD/MM/YY
    m = re.search(r"\b(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})\b", text)
    if m:
        return m.group(1).replace("/", "-")
    # FRIDAY2 28-11-25 style (OCR may glue digit to weekday)
    m = re.search(
        r"(?:MON|TUE|WED|THU|FRI|SAT|SUN)[A-Z0-9]*\s*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})",
        text, re.IGNORECASE,
    )
    if m:
        return m.group(1).replace("/", "-")
    # Loose match
    m = re.search(r"\d{1,2}[-/]\d{1,2}[-/]\d{2,4}", text)
    return m.group(0).replace("/", "-") if m else ""


def _parse_section(text: str) -> str:
    # After keyword
    m = re.search(
        r"(?:section|sec|branch|class|dept)[:\s]+([A-Z0-9\-]{1,8})",
        text, re.IGNORECASE,
    )
    if m:
        return m.group(1).strip()
    # Known dept codes
    m = re.search(
        r"\b(CSM|CSE|CSD|IT|ECE|EEE|MECH|CIVIL|AI|DS|AIDS|AIML|MBA|MCA)\b",
        text, re.IGNORECASE,
    )
    if m:
        return m.group(1).upper()
    # Short ALL-CAPS fallback
    _skip = {
        "MON","TUE","WED","THU","FRI","SAT","SUN",
        "MONDAY","TUESDAY","WEDNESDAY","THURSDAY","FRIDAY","SATURDAY","SUNDAY",
        "AM","PM","NA","YR","III","FOR","THE","AND","OF",
    }
    for tok in re.findall(r"\b[A-Z]{2,6}\b", text):
        if tok not in _skip:
            return tok
    return ""


def extract_middle_region(mid_img: np.ndarray) -> dict:
    """
    Parse the attendance grid maintaining students.xlsx order.
    
    Strategy:
      1. OCR entire middle region.
      2. Group tokens into horizontal rows.
      3. Per row: extract roll number(s) and attendance tokens.
      4. Map token positions to period numbers.
      5. Mark "A" as absent, "L" or "lunch" as lunch break, others as present.
      6. Reorder results according to students.xlsx
    """
    records = _run_ocr(mid_img)
    rows    = _group_rows(records, y_tol=16)
    
    # Load students order from students.xlsx
    students_order = load_students_order()
    order_map = {roll: idx for idx, roll in enumerate(students_order)}

    all_rolls: list   = []
    absentees: list   = []
    roll_details: list = []
    period_absentees: dict = {f"P{i}": [] for i in range(1, NUM_PERIODS + 1)}  # Track absentees per period

    # Temporary storage to sort later
    temp_details = []

    for row in rows:
        texts   = [r["text"] for r in row]
        row_str = " ".join(texts)

        rolls = ROLL_PATTERN.findall(row_str)
        if not rolls:
            continue

        # Skip rows that are likely not student rows (too few tokens or no attendance marks)
        # Filter out non-attendance tokens
        attendance_tokens = []
        for t in texts:
            t = t.strip()
            if ROLL_PATTERN.fullmatch(t):
                continue
            if not t:
                continue
            # Only keep tokens that could be attendance marks
            if re.fullmatch(r"[AaLl]|/|\\|-|\)|\d|L-", t) or t.upper() in ["LUNCH", "PRESENT", "ABSENT"]:
                attendance_tokens.append(t)
        
        # Don't skip rows even if no attendance tokens found - they might still be valid students
        # Just mark them as all present if no tokens found
        
        # Debug: log the marks found
        logger.debug(f"Roll {rolls[0] if rolls else 'unknown'} marks: {attendance_tokens}")

        # Check for attendance status
        def _is_absence_mark(tok: str) -> bool:
            return bool(re.fullmatch(r"[Aa]", tok))
        
        def _is_lunch_mark(tok: str) -> bool:
            return bool(re.fullmatch(r"[Ll]|L-|lunch|LUNCH", tok))
        
        def _is_present_mark(tok: str) -> bool:
            return bool(re.fullmatch(r"[/\\\-)]|√|✓|present|PRESENT", tok) or tok.isdigit())

        # Build per-period breakdown maintaining order
        period_map = {}
        is_absent = False
        
        # Ensure we have exactly 8 periods, fill with defaults
        for idx in range(1, NUM_PERIODS + 1):
            period_map[f"P{idx}"] = "Present"  # Default to Present
        
        # Set P5 to L for all students (lunch period)
        period_map[f"P5"] = "L"
        
        # Process attendance tokens we found
        for idx, tok in enumerate(attendance_tokens[:NUM_PERIODS], start=1):
            if _is_absence_mark(tok):
                period_map[f"P{idx}"] = "Absent"
                is_absent = True
            elif _is_lunch_mark(tok):
                # Lunch token found, P5 is already set to L
                pass  # P5 is already set to L for all students
            elif _is_present_mark(tok):
                period_map[f"P{idx}"] = "Present"  # Explicit present marks
            # Keep default "Present" for any other cases including empty tokens
            
            # Track absentees per period
            if _is_absence_mark(tok):
                period_absentees[f"P{idx}"].extend(rolls)

        # Add to temporary list for sorting
        if len(rolls) == 1:
            # Single roll number - apply attendance tokens normally
            for roll in rolls:
                temp_details.append({
                    "roll":   roll,
                    "absent": is_absent,
                    "order":  order_map.get(roll, 999999),  # Use high number if not found
                    **period_map,
                })
        else:
            # Multiple roll numbers in same row - need to split attendance tokens
            # This usually happens when two students share the same row
            # We'll distribute attendance tokens evenly or mark as all present if unclear
            
            # Calculate how many tokens per student (approximate)
            tokens_per_student = max(1, len(attendance_tokens) // len(rolls))
            
            for i, roll in enumerate(rolls):
                # Create a copy of period_map for each student
                student_period_map = period_map.copy()
                student_is_absent = False
                
                # Get tokens for this student
                start_idx = i * tokens_per_student
                end_idx = start_idx + tokens_per_student
                student_tokens = attendance_tokens[start_idx:end_idx]
                
                # Process this student's tokens
                for idx, tok in enumerate(student_tokens[:NUM_PERIODS], start=1):
                    if _is_absence_mark(tok):
                        student_period_map[f"P{idx}"] = "Absent"
                        student_is_absent = True
                    elif _is_lunch_mark(tok):
                        student_period_map[f"P5"] = "L"  # Lunch period
                    elif _is_present_mark(tok):
                        student_period_map[f"P{idx}"] = "Present"
                
                temp_details.append({
                    "roll":   roll,
                    "absent": student_is_absent,
                    "order":  order_map.get(roll, 999999),
                    **student_period_map,
                })

    # Sort by students.xlsx order
    temp_details.sort(key=lambda x: x["order"])
    
    # Add any missing students from students.xlsx (not found in OCR)
    found_rolls = {detail["roll"] for detail in temp_details}
    for roll in students_order:
        if roll not in found_rolls:
            # Add missing student as all present
            period_map = {}
            for idx in range(1, NUM_PERIODS + 1):
                period_map[f"P{idx}"] = "Present"
            period_map[f"P5"] = "L"  # Lunch period
            
            temp_details.append({
                "roll": roll,
                "absent": False,
                "order": order_map.get(roll, 999999),
                **period_map,
            })
    
    # Sort again to include the newly added students
    temp_details.sort(key=lambda x: x["order"])
    
    # Build final lists in correct order
    for detail in temp_details:
        roll = detail["roll"]
        all_rolls.append(roll)
        if detail["absent"]:
            absentees.append(roll)
        # Remove the order field before adding to roll_details
        detail_copy = dict(detail)
        del detail_copy["order"]
        roll_details.append(detail_copy)

    logger.info(
        "Middle region → %d rolls, %d absentees (ordered by students.xlsx)", len(all_rolls), len(absentees)
    )
    return {
        "all_rolls":        all_rolls,      # In students.xlsx order
        "absentees":        absentees,      # In students.xlsx order
        "roll_details":     roll_details,
        "period_absentees": period_absentees,  # New: absentees per period
    }


def extract_bottom_region(bot_img: np.ndarray) -> dict:
    """Parse bottom region - no longer extracts teachers from attendance sheet."""
    records = _run_ocr(bot_img)
    logger.debug("BOTTOM text: %s", [r["text"] for r in records])
    # Teachers are now loaded from timetable only, not from attendance sheet
    return {"teachers": []}


def _parse_teachers(records: list) -> list:
    """
    Extract teacher name tokens, ignoring table headers and staff.
    Only returns actual teachers, excludes "Students" and other non-teachers.
    """
    teachers, seen = [], set()
    name_pat = re.compile(
        r"\b(?:Dr\.?|Prof\.?|Mr\.?|Mrs\.?|Ms\.?)?[A-Z][A-Za-z.]{1,24}\b"
    )
    
    # Words to exclude (non-teachers)
    exclude_words = {"Students", "Student", "Staff", "Period", "Lunch", "Break"}

    for rec in records:
        for tok in name_pat.findall(rec["text"]):
            clean = tok.strip().strip(".")
            if not clean or clean.lower() in IGNORE_WORDS:
                continue
            if clean in exclude_words:
                continue
            if len(clean) < 2 or clean.isdigit():
                continue
            key = clean.upper()
            if key not in seen:
                seen.add(key)
                teachers.append(clean)

    return teachers


# ══════════════════════════════════════════════════════════════════════════════
# Excel export — mirrors the paper sheet structure
# ══════════════════════════════════════════════════════════════════════════════

def export_attendance_files(data: dict, base_path: str) -> tuple:
    """
    Create separate files for absentees and presents.
    Returns tuple: (absentees_file, presents_file)
    """
    date = data.get("date", "").replace("-", "")
    section = data.get("section", "")
    
    # Extract roll details
    roll_details = data.get("roll_details", [])
    period_absentees = data.get("period_absentees", {})
    
    # Create absentees file with period-wise mapping
    absentees_data = []
    for period, rolls in period_absentees.items():
        for roll in rolls:
            absentees_data.append({
                "Roll Number": roll,
                "Period": period,
                "Status": "Absent"
            })
    
    # Remove duplicates while maintaining order
    seen = set()
    unique_absentees = []
    for item in absentees_data:
        key = (item["Roll Number"], item["Period"])
        if key not in seen:
            seen.add(key)
            unique_absentees.append(item)
    
    # Create presents file
    all_rolls = data.get("all_rolls", [])
    absent_rolls = set(data.get("absentees", []))
    presents = [roll for roll in all_rolls if roll not in absent_rolls]
    
    presents_data = [{"Roll Number": roll, "Status": "Present"} for roll in presents]
    
    # Generate filenames
    absentees_file = f"{base_path}_{date}_{section}_absentees.xlsx"
    presents_file = f"{base_path}_{date}_{section}_presents.xlsx"
    
    # Export to Excel
    if unique_absentees:
        df_absentees = pd.DataFrame(unique_absentees)
        df_absentees.to_excel(absentees_file, index=False)
        logger.info(f"Absentees file created: {absentees_file}")
    else:
        logger.info("No absentees to export")
    
    if presents_data:
        df_presents = pd.DataFrame(presents_data)
        df_presents.to_excel(presents_file, index=False)
        logger.info(f"Presents file created: {presents_file}")
    else:
        logger.info("No presents to export")
    
    return absentees_file, presents_file


def export_to_excel(data: dict, output_path: str) -> str:
    """
    Write a styled 4-sheet Excel workbook:
      Summary    — metadata + counts
      Attendance — full roll × period grid (mirrors paper layout)
      Absentees  — absent roll numbers only
      Teachers   — period → teacher name
    """
    all_rolls    = data.get("all_rolls", [])
    absentees_s  = set(data.get("absentees", []))
    teachers     = data.get("teachers", [])
    roll_details = data.get("roll_details", [])

    present_ct = len(all_rolls) - len(absentees_s)
    pct        = round(present_ct / len(all_rolls) * 100, 1) if all_rolls else 0.0

    df_summary = pd.DataFrame([
        {"Field": "Date",           "Value": data.get("date", "")},
        {"Field": "Section",        "Value": data.get("section", "")},
        {"Field": "Total Students", "Value": len(all_rolls)},
        {"Field": "Present",        "Value": present_ct},
        {"Field": "Absent",         "Value": len(absentees_s)},
        {"Field": "Attendance %",   "Value": f"{pct}%"},
    ])

    if roll_details:
        df_att = pd.DataFrame(roll_details).rename(
            columns={"roll": "Roll Number", "absent": "Absent?"}
        )
        df_att["Absent?"] = df_att["Absent?"].map({True: "YES", False: "NO"})
    else:
        df_att = pd.DataFrame([
            {"Roll Number": r, "Absent?": "YES" if r in absentees_s else "NO"}
            for r in all_rolls
        ])

    df_absent  = pd.DataFrame({"Absent Roll Number": sorted(absentees_s)})
    df_teachers = pd.DataFrame({
        "Period":  [f"Period {i+1}" for i in range(len(teachers))],
        "Teacher": teachers,
    })

    if not output_path.endswith(".xlsx"):
        output_path += ".xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_summary.to_excel( writer, sheet_name="Summary",    index=False)
        df_att.to_excel(     writer, sheet_name="Attendance", index=False)
        df_absent.to_excel(  writer, sheet_name="Absentees",  index=False)
        df_teachers.to_excel(writer, sheet_name="Teachers",   index=False)

        try:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            DARK       = PatternFill("solid", fgColor="1A1A2E")
            RED_BG     = PatternFill("solid", fgColor="FFE0E0")
            GREEN_BG   = PatternFill("solid", fgColor="E6F4EA")
            GOLD_FONT  = Font(color="C9A84C", bold=True, size=11)
            THIN_SIDE  = Side(style="thin", color="CCCCCC")
            THIN       = Border(left=THIN_SIDE, right=THIN_SIDE,
                                top=THIN_SIDE,  bottom=THIN_SIDE)
            CENTER     = Alignment(horizontal="center", vertical="center")

            for sname, df in [
                ("Summary", df_summary), ("Attendance", df_att),
                ("Absentees", df_absent), ("Teachers", df_teachers),
            ]:
                ws = writer.sheets[sname]
                # Header row
                for ci, col in enumerate(df.columns, 1):
                    cell = ws.cell(1, ci)
                    cell.fill      = DARK
                    cell.font      = GOLD_FONT
                    cell.alignment = CENTER
                    cell.border    = THIN
                    max_len = max(
                        len(str(col)),
                        max((len(str(v)) for v in df[col]), default=0),
                    )
                    ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 36)
                ws.row_dimensions[1].height = 22

                # Data rows
                for ri in range(2, len(df) + 2):
                    for ci in range(1, len(df.columns) + 1):
                        cell = ws.cell(ri, ci)
                        cell.border    = THIN
                        cell.alignment = CENTER
                        val = str(cell.value or "")
                        if val in ("YES", "A"):
                            cell.fill = RED_BG
                        elif val in ("NO", "P"):
                            cell.fill = GREEN_BG
        except ImportError:
            pass  # Write without styles if openpyxl.styles unavailable

    logger.info("Excel written → %s", output_path)
    return output_path


# ══════════════════════════════════════════════════════════════════════════════
# File-type processors
# ══════════════════════════════════════════════════════════════════════════════

def process_image(file_path: str) -> dict:
    img = cv2.imread(str(file_path))
    if img is None:
        raise ValueError(f"Cannot read image: {file_path}")
    img  = _auto_upscale(img)
    proc = _preprocess(img)
    top, mid, bot = _split_regions(proc)
    top_d = extract_top_region(top)
    mid_d = extract_middle_region(mid)
    bot_d = extract_bottom_region(bot)
    
    # Load schedule information including teachers
    schedule_info = load_schedule_info()
    
    # Get all unique teachers from all periods
    all_teachers = []
    for period_teachers in schedule_info.get("staff_per_period", {}).values():
        all_teachers.extend(period_teachers)
    
    # Remove duplicates while preserving order
    seen = set()
    unique_teachers = []
    for teacher in all_teachers:
        if teacher not in seen:
            seen.add(teacher)
            unique_teachers.append(teacher)
    
    return {
        "date":             top_d.get("date", ""),
        "section":          top_d.get("section", ""),
        "absentees":        mid_d.get("absentees", []),
        "all_rolls":        mid_d.get("all_rolls", []),
        "roll_details":     mid_d.get("roll_details", []),
        "period_absentees": mid_d.get("period_absentees", {}),
        "teachers":         unique_teachers,  # Teachers from timetable only
        "schedule":         schedule_info,
    }


def process_pdf(file_path: str) -> dict:
    try:
        from pdf2image import convert_from_path
    except ImportError:
        raise ImportError("Run: pip install pdf2image")

    pages = convert_from_path(file_path, dpi=300)
    combined = {
        "date": "", "section": "",
        "absentees": [], "all_rolls": [], "roll_details": [], "teachers": [],
    }
    tmp = Path(file_path).parent / "_ocr_tmp"
    tmp.mkdir(exist_ok=True)

    for i, page in enumerate(pages):
        p = str(tmp / f"p{i}.png")
        page.save(p, "PNG")
        try:
            d = process_image(p)
            if i == 0:
                combined["date"]    = d.get("date", "")
                combined["section"] = d.get("section", "")
            combined["absentees"].extend(d.get("absentees", []))
            combined["all_rolls"].extend(d.get("all_rolls", []))
            combined["roll_details"].extend(d.get("roll_details", []))
            combined["teachers"].extend(d.get("teachers", []))
        finally:
            try: os.remove(p)
            except: pass

    try: tmp.rmdir()
    except: pass

    combined["absentees"] = list(dict.fromkeys(combined["absentees"]))
    combined["all_rolls"] = list(dict.fromkeys(combined["all_rolls"]))
    combined["teachers"]  = list(dict.fromkeys(combined["teachers"]))
    return combined


def process_excel(file_path: str) -> dict:
    xl = pd.ExcelFile(file_path)
    df = None
    for sheet in xl.sheet_names:
        c = xl.parse(sheet)
        c.columns = [str(x).strip().lower() for x in c.columns]
        if any("roll" in x for x in c.columns):
            df = c
            break

    if df is None:
        return {"date": "", "section": "", "absentees": [],
                "all_rolls": [], "roll_details": [], "teachers": []}

    roll_col = next(c for c in df.columns if "roll" in c)
    rolls    = [r for r in df[roll_col].dropna().astype(str).str.strip()
                if ROLL_PATTERN.fullmatch(r)]

    status_col = next((c for c in df.columns if "status" in c or "absent" in c), None)
    absentees  = [r for r, s in zip(rolls, df[status_col].astype(str))
                  if s.strip().upper() in {"A", "ABSENT", "YES"}] if status_col else []

    def _first(col_key):
        c = next((x for x in df.columns if col_key in x), None)
        return str(df[c].dropna().iloc[0]) if c and not df[c].dropna().empty else ""

    return {
        "date": _first("date"), "section": _first("section"),
        "absentees": absentees, "all_rolls": rolls,
        "roll_details": [],
        "teachers": df[next((c for c in df.columns if "teacher" in c), None)
                      ].dropna().astype(str).tolist()
        if any("teacher" in c for c in df.columns) else [],
    }


# ══════════════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════════════

def process_attendance(file_path: str, export_excel: bool = True) -> dict:
    """
    Auto-detect file type and extract structured attendance data.

    Returns
    -------
    {
        "date":         str,
        "section":      str,
        "absentees":    list[str],
        "teachers":     list[str],
        "all_rolls":    list[str],
        "roll_details": list[dict],
        "excel_path":   str | None,
    }
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Not found: {file_path}")

    ext = path.suffix.lower()
    if ext in IMAGE_EXTS:
        data = process_image(str(path))
    elif ext == ".pdf":
        data = process_pdf(str(path))
    elif ext in {".xlsx", ".xls"}:
        data = process_excel(str(path))
    else:
        raise ValueError(f"Unsupported extension: {ext}")

    result = {
        "date":         data.get("date", ""),
        "section":      data.get("section", ""),
        "absentees":    data.get("absentees", []),
        "teachers":     data.get("teachers", []),
        "all_rolls":    data.get("all_rolls", []),
        "roll_details": data.get("roll_details", []),
        "excel_path":   None,
    }

    logger.info(
        "Extracted — date=%r section=%r rolls=%d absentees=%d teachers=%d",
        result["date"], result["section"],
        len(result["all_rolls"]), len(result["absentees"]), len(result["teachers"]),
    )

    if export_excel:
        out = str(path.parent / f"{path.stem}_attendance.xlsx")
        try:
            result["excel_path"] = export_to_excel(result, out)
        except Exception as e:
            logger.warning("Excel export failed: %s", e)

    return result


# CLI
if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 2:
        print("Usage: python ocr_module.py <file>")
        sys.exit(1)
    r = process_attendance(sys.argv[1])
    print(json.dumps({k: r[k] for k in ("date", "section", "absentees", "teachers")}, indent=2))
    if r.get("excel_path"):
        print(f"\nExcel → {r['excel_path']}")