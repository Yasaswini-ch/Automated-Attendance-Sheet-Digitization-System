import openpyxl
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill,
    GradientFill
)
from openpyxl.utils import get_column_letter
import io


# Styles
THIN = Side(border_style="thin", color="000000")
MEDIUM = Side(border_style="medium", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MEDIUM_BORDER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
LUNCH_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
UNCLEAR_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
BOLD_FONT = Font(bold=True, size=10)
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=12)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_border(cell, border=THIN_BORDER):
    cell.border = border


def _set_cell(ws, row, col, value, font=None, alignment=None, fill=None, border=THIN_BORDER):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    return cell


def generate_excel(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Sheet"

    header = data.get("header", {})
    attendance = data.get("attendance_table", [])
    bottom = data.get("bottom_section", {})

    # ---- ROW 1: College Name ----
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = header.get("college_name", "")
    cell.font = Font(bold=True, size=14)
    cell.alignment = CENTER
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER

    # ---- ROW 2: Location ----
    ws.merge_cells("A2:L2")
    cell = ws["A2"]
    cell.value = header.get("location", "")
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER

    # ---- ROW 3: Day/Date | Department | Year/Semester | Section ----
    # Day & Date label
    ws.merge_cells("A3:B3")
    _set_cell(ws, 3, 1, "DAY & DATE", BOLD_FONT, CENTER, SUBHEADER_FILL)
    ws.merge_cells("C3:F3")
    _set_cell(ws, 3, 3, "DEPARTMENT / BRANCH", BOLD_FONT, CENTER, SUBHEADER_FILL)
    ws.merge_cells("G3:I3")
    _set_cell(ws, 3, 7, "YEAR / SEMESTER", BOLD_FONT, CENTER, SUBHEADER_FILL)
    ws.merge_cells("J3:K3")
    _set_cell(ws, 3, 10, "Section", BOLD_FONT, CENTER, SUBHEADER_FILL)

    # ---- ROW 4: Values ----
    ws.merge_cells("A4:B4")
    _set_cell(ws, 4, 1, header.get("day_date", ""), None, CENTER)
    ws.merge_cells("C4:F4")
    _set_cell(ws, 4, 3, header.get("department_branch", ""), None, CENTER)
    ws.merge_cells("G4:I4")
    _set_cell(ws, 4, 7, header.get("year_semester", ""), None, CENTER)
    ws.merge_cells("J4:K4")
    _set_cell(ws, 4, 10, header.get("section", ""), None, CENTER)

    # Apply borders to row 3&4
    for col in range(1, 12):
        ws.cell(row=3, column=col).border = THIN_BORDER
        ws.cell(row=4, column=col).border = THIN_BORDER

    # ---- ROW 5: "Period" spanning header ----
    ws.merge_cells("A5:A6")
    _set_cell(ws, 5, 1, "H.T.No", BOLD_FONT, CENTER, SUBHEADER_FILL)
    
    ws.merge_cells("B5:L5")
    cell = ws.cell(row=5, column=2, value="Period")
    cell.font = BOLD_FONT
    cell.alignment = CENTER
    cell.fill = SUBHEADER_FILL
    cell.border = THIN_BORDER

    # ---- ROW 6: Period numbers ----
    period_cols = {
        "1": 2, "2": 3, "3": 4, "4": 5,
        "LUNCH": 6,
        "5": 7, "6": 8, "7": 9, "8": 10
    }
    
    period_labels = ["1", "2", "3", "4", "LUNCH", "5", "6", "7", "8"]
    for i, label in enumerate(period_labels):
        col = i + 2
        fill = LUNCH_FILL if label == "LUNCH" else SUBHEADER_FILL
        _set_cell(ws, 6, col, label, BOLD_FONT, CENTER, fill)

    # ---- Attendance Rows ----
    START_ROW = 7
    for i, row_data in enumerate(attendance):
        excel_row = START_ROW + i
        ht_no = row_data.get("ht_no", "")
        
        # HT No - left aligned
        cell = ws.cell(row=excel_row, column=1, value=ht_no)
        cell.font = Font(size=9)
        cell.alignment = LEFT
        cell.border = THIN_BORDER

        # Period values
        period_keys = ["period1", "period2", "period3", "period4", "lunch", "period5", "period6", "period7", "period8"]
        for j, key in enumerate(period_keys):
            col = j + 2
            val = row_data.get(key, "")
            
            fill = LUNCH_FILL if key == "lunch" else None
            if str(val).strip().upper() == "UNCLEAR":
                fill = UNCLEAR_FILL
            
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.font = Font(size=9)
            if fill:
                cell.fill = fill

    # ---- Bottom Section ----
    BOTTOM_START = START_ROW + len(attendance) + 1
    
    teacher_names = bottom.get("teacher_names", {})
    subjects = bottom.get("subjects", {})
    
    bottom_rows = [
        ("Period", ["", "1", "2", "3", "4", "LUNCH", "5", "6", "7", "8"]),
        ("Name of the Teacher", ["", 
            teacher_names.get("period1",""), teacher_names.get("period2",""),
            teacher_names.get("period3",""), teacher_names.get("period4",""),
            "LUNCH",
            teacher_names.get("period5",""), teacher_names.get("period6",""),
            teacher_names.get("period7",""), teacher_names.get("period8","")
        ]),
        ("Signature of the Teacher", [""] * 10),
        ("Substitute if any", [""] * 10),
        ("Time table for the Day", [""] * 10),
        ("Total Students", ["", "No. Present", "", "", "", "", "", "No. Absent", "", ""]),
    ]
    
    # Subject row
    subject_row_data = ["",
        subjects.get("period1",""), subjects.get("period2",""),
        subjects.get("period3",""), subjects.get("period4",""),
        "",
        subjects.get("period5",""), subjects.get("period6",""),
        subjects.get("period7",""), subjects.get("period8","")
    ]

    for r_offset, (label, values) in enumerate(bottom_rows):
        excel_row = BOTTOM_START + r_offset
        # Label cell
        cell = ws.cell(row=excel_row, column=1, value=label)
        cell.font = BOLD_FONT
        cell.alignment = LEFT
        cell.border = THIN_BORDER
        cell.fill = SUBHEADER_FILL
        
        for c_offset, val in enumerate(values):
            col = c_offset + 1
            if col == 1:
                continue
            fill = LUNCH_FILL if val == "LUNCH" else None
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.font = Font(size=9)
            if fill:
                cell.fill = fill
    
    # Insert subject row after teacher name row
    subject_excel_row = BOTTOM_START + 1  # same as teacher name row, append below
    # Actually insert as sub-row by adjusting
    # We'll add it as extra info in teacher cell
    # Better: add a dedicated subject row
    ws.insert_rows(BOTTOM_START + 2)
    subject_excel_row = BOTTOM_START + 2
    cell = ws.cell(row=subject_excel_row, column=1, value="Subject")
    cell.font = BOLD_FONT
    cell.alignment = LEFT
    cell.border = THIN_BORDER
    cell.fill = SUBHEADER_FILL
    
    for c_offset, val in enumerate(subject_row_data):
        col = c_offset + 1
        if col == 1:
            continue
        cell = ws.cell(row=subject_excel_row, column=col, value=val)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.font = Font(size=9)

    # ---- Column Widths ----
    ws.column_dimensions["A"].width = 18  # HT No
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col_letter].width = 8
    ws.column_dimensions["F"].width = 7  # LUNCH

    # Row heights
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16
    for r in range(3, 7):
        ws.row_dimensions[r].height = 18
    for r in range(START_ROW, START_ROW + len(attendance)):
        ws.row_dimensions[r].height = 15

    # ---- Legend ----
    legend_row = BOTTOM_START + 10
    ws.cell(row=legend_row, column=1, value="Legend:").font = BOLD_FONT
    ws.cell(row=legend_row, column=2, value="UNCLEAR").fill = UNCLEAR_FILL
    ws.cell(row=legend_row, column=3, value="= Could not read value")
    ws.cell(row=legend_row+1, column=2, value="LUNCH").fill = LUNCH_FILL
    ws.cell(row=legend_row+1, column=3, value="= Lunch break separator")

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
